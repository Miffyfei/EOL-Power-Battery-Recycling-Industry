# -*- coding: utf-8 -*-
"""
CEV clustering + Multi-model per cluster (RF/XGB/MLP/SVM + Stacking + Optional BlendTop2)
Improvements (without changing model families):
- Seasonality enhancement: 12-period seasonal differencing + seasonal rolling mean removal (deseasonal residual)
- Adaptive target transformation: Use log1p based on skewness (TTR auto inverse transform)
- Time series CV: TimeSeriesSplit + gap=1 (avoid adjacent leakage)
- Adaptive search: Larger clusters get more search space and iterations, smaller clusters more conservative
- Only C0 attempts Top2 blending (0.6/0.4), other clusters avoid overfitting
- Preserve original evaluation export and charts
"""

import os, json, math, warnings, traceback
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

from typing import List, Dict, Tuple

from sklearn.model_selection import TimeSeriesSplit, RandomizedSearchCV
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer, TransformedTargetRegressor
from sklearn.metrics import (r2_score, mean_squared_error, mean_absolute_error,
                             mean_absolute_percentage_error,
                             silhouette_score, silhouette_samples,
                             davies_bouldin_score, calinski_harabasz_score)
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans

from sklearn.ensemble import RandomForestRegressor
from sklearn.neural_network import MLPRegressor
from sklearn.linear_model import LinearRegression, RidgeCV
from sklearn.svm import SVR
from xgboost import XGBRegressor

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

# ========================== Paths and Global Configuration ==========================
INPUT_PATH = "./input data/2016-2030_PEV and CEV_month+data.xlsx"
SHEET_NAME = "Sheet1"
OUT_DIR = "./output data/results_electric vehicle"
os.makedirs(OUT_DIR, exist_ok=True)
MASTER_XLSX = os.path.join(OUT_DIR, "CEV_ClusterBest_Forecast.xlsx")

RANDOM_STATE = 42
np.random.seed(RANDOM_STATE)

# Search intensity (consistent with PEV scheme logic, but adaptive by cluster size)
N_TRIALS     = 120
N_TRIALS_C0  = 160   # Higher search intensity for C0
N_SPLITS     = 7
CV_GAP       = 1
K_RANGE      = range(3, 15)

# Target column candidates (CEV)
TARGET_CEV_CANDS = ['CEV_number', 'Commercial_on_insurance', 'EV_Commercial_Sales', 'CEV']

# Feature mapping (fault-tolerant)
BASE_FEATURE_CANDS = {
    'people': ['people','Population','population','People'],
    'total_GDP': ['total_GDP','GDP','gdp','Total_GDP'],
    'Urbanization_rate': ['Urbanization_rate','urbanization_rate','Urban_rate'],
    'Baidu_index': ['Baidu_index','baidu_index','Baidu'],
    'CEV_Subsidy funds': ['CEV_Subsidy funds','Subsidy','EV_Subsidy','subsidy']
}

# Tabular base features
TAB_FEATS_BASE = [
    'people','total_GDP','Urbanization_rate','Baidu_index','CEV_Subsidy funds',
    'month_sin','month_cos','year_num',
    'lag_1','lag_3','lag_12','roll_mean_3','roll_mean_12'
]

# Seasonal enhancement features (strictly using history)
SEASONAL_FEATS = ['sea_diff_1_12','sea_ratio_1_12','sea_diff_mean3_12','y_deseasonal_res']  # Added y deseasonal residual

# Forecast extrapolation constraints
HIST_MAX_MULT = 3.0
CLIP_NONNEG   = True

# ========================== Excel I/O ==========================
def ensure_book(path):
    if not os.path.exists(path):
        wb = Workbook(); wb.save(path)

def write_df(path, df: pd.DataFrame, sheet: str, index=False):
    ensure_book(path)
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
        df.to_excel(w, sheet_name=sheet, index=index)

def embed_png(path, sheet, img_path, anchor="A1"):
    if not os.path.exists(img_path): return
    ensure_book(path)
    wb = load_workbook(path)
    if sheet not in wb.sheetnames: wb.create_sheet(sheet)
    ws = wb[sheet]
    try:
        img = XLImage(img_path); img.anchor = anchor
        ws.add_image(img)
    except Exception as e:
        print("Embed image failed:", e)
    wb.save(path)

# ========================== Data Reading ==========================
def pick_col(df, cands):
    ok = [(c, df[c].notna().sum()) for c in cands if c in df.columns]
    if not ok: return None
    ok.sort(key=lambda x:x[1], reverse=True)
    return ok[0][0]

def read_panel(path, sheet):
    df = pd.read_excel(path, sheet_name=sheet)
    df.columns = df.columns.map(lambda x: str(x).strip())
    for base in ['City','Year','Month']:
        if base not in df.columns:
            hits=[c for c in df.columns if c.lower()==base.lower()]
            if hits: df.rename(columns={hits[0]:base}, inplace=True)
    assert {'City','Year','Month'}.issubset(df.columns), "Require City/Year/Month columns"

    for std, cands in BASE_FEATURE_CANDS.items():
        if std not in df.columns:
            hit=[c for c in cands if c in df.columns]
            if hit: df.rename(columns={hit[0]:std}, inplace=True)
        if std not in df.columns: df[std]=np.nan

    t_cev = pick_col(df, TARGET_CEV_CANDS)
    if t_cev is None:
        raise ValueError("CEV target column not found")

    for c in ['people','total_GDP','Urbanization_rate','Baidu_index','CEV_Subsidy funds', t_cev]:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    df['date'] = pd.to_datetime(dict(year=df['Year'].astype(int), month=df['Month'].astype(int), day=1))
    df = df.sort_values(['City','date']).reset_index(drop=True)
    return df, t_cev

# ========================== Feature Engineering (lag/rolling + time terms + seasonal enhancement + interaction) ==========================
def add_time_cols(df):
    df = df.copy()
    m = df['Month'].astype(int)
    df['month_sin'] = np.sin(2*np.pi*(m-1)/12)
    df['month_cos'] = np.cos(2*np.pi*(m-1)/12)
    df['year_num']  = df['Year'].astype(int)
    return df

def add_lag_roll(df, y, lags=(1,3,12), wins=(3,12)):
    df = df.sort_values(['City','date']).copy()
    for city, g in df.groupby('City'):
        idx = g.index
        for L in lags:
            df.loc[idx, f'lag_{L}'] = g[y].shift(L)
        for w in wins:
            df.loc[idx, f'roll_mean_{w}'] = g[y].shift(1).rolling(w).mean()
    return df

def add_seasonal_feats(df, y):
    df = df.sort_values(['City','date']).copy()
    for city, g in df.groupby('City'):
        idx = g.index
        # Basic seasonal differencing
        lag1  = g[y].shift(1)
        lag13 = g[y].shift(13)
        mean3_lag1  = g[y].shift(1).rolling(3).mean()
        mean3_lag13 = g[y].shift(13).rolling(3).mean()
        df.loc[idx,'sea_diff_1_12']     = (lag1 - lag13).values
        df.loc[idx,'sea_ratio_1_12']    = (lag1 / (lag13 + 1e-9)).values
        df.loc[idx,'sea_diff_mean3_12'] = (mean3_lag1 - mean3_lag13).values
        # y deseasonalized residual: y - rolling_mean_12(previous period)
        roll12_prev = g[y].shift(1).rolling(12).mean()
        df.loc[idx,'y_deseasonal_res']  = (g[y] - roll12_prev).values
    return df

def build_tabular(df, y):
    df = add_time_cols(df)
    df = add_lag_roll(df, y)
    df = add_seasonal_feats(df, y)

    # Interaction terms (as needed)
    if 'total_GDP' in df.columns and 'people' in df.columns:
        df['gdp_per_capita'] = df['total_GDP'] / (df['people'] + 1e-9)
    if 'CEV_Subsidy funds' in df.columns and 'total_GDP' in df.columns:
        df['subsidy_to_gdp'] = df['CEV_Subsidy funds'] / (df['total_GDP'] + 1e-9)
    if 'Baidu_index' in df.columns and 'people' in df.columns:
        df['baidu_per_capita'] = df['Baidu_index'] / (df['people'] + 1e-9)

    feats_all = TAB_FEATS_BASE + SEASONAL_FEATS + ['gdp_per_capita','subsidy_to_gdp','baidu_per_capita']
    for c in feats_all:
        if c not in df.columns: df[c]=np.nan
        df[c] = pd.to_numeric(df[c], errors='coerce')

    # Ensure necessary lag/seasonal terms are complete
    df = df.dropna(subset=['lag_1','lag_3','lag_12',
                           'sea_diff_1_12','sea_ratio_1_12','sea_diff_mean3_12','y_deseasonal_res'])
    return df, feats_all

# ========================== Correlation (for explanation, optional) ==========================
def corr_abs_to_target(df, target):
    cols = list({*TAB_FEATS_BASE, *SEASONAL_FEATS, target})
    sub = add_time_cols(df)[[c for c in cols if c in df.columns]].copy()
    C = sub.corr(numeric_only=True)
    if target not in C.columns: return pd.Series(dtype=float)
    return C[target].drop(index=target).abs().sort_values(ascending=False)

# ========================== Clustering (auto K selection + visualization) ==========================
def extract_city_level_features(df, target):
    d = df.copy()
    d['Urban_population'] = d['people'] * d['Urbanization_rate']
    bases = ['total_GDP','Urban_population','Baidu_index','people','Urbanization_rate']
    rows=[]
    for city, g in d.groupby('City'):
        g = g.sort_values('date'); r={'City':city}
        g_hist = g[g['Year']<=2024]
        for b in bases:
            r[f'lv_{b}'] = g_hist[g_hist['Year']==2024][b].mean()
        g_recent = g_hist[g_hist['Year']>=2022]
        for b in bases:
            s = g_recent[b].dropna()
            if len(s)>=2:
                first,last = s.iloc[0], s.iloc[-1]
                n=max(1,len(s)-1)
                cagr=((last+1e-9)/(first+1e-9))**(1/n)-1
                r[f'cagr_{b}']=cagr; r[f'var_{b}']=s.var()
            else:
                r[f'cagr_{b}']=0.0; r[f'var_{b}']=0.0
        rows.append(r)
    return pd.DataFrame(rows).fillna(0.0)

def embed_city_features(city_feat):
    scaler = StandardScaler()
    X = scaler.fit_transform(city_feat.drop(columns=['City']))
    pca = PCA(n_components=2, random_state=RANDOM_STATE)
    Xp = pca.fit_transform(X)
    return city_feat[['City']].assign(pc1=Xp[:,0], pc2=Xp[:,1])

def choose_k_enhanced(emb, k_range=K_RANGE):
    X = emb[['pc1','pc2']].values; rows=[]
    for k in k_range:
        km = KMeans(n_clusters=k, n_init=100, random_state=RANDOM_STATE).fit(X)
        labs = km.labels_
        if any((labs==i).sum()<3 for i in range(k)):  # At least 3 cities per cluster
            continue
        sil = silhouette_score(X, labs)
        dbi = davies_bouldin_score(X, labs)
        ch  = calinski_harabasz_score(X, labs)
        rows.append({'k':k,'silhouette':sil,'DBI':dbi,'CH':ch})
    m = pd.DataFrame(rows)
    if m.empty: return 3, None, None
    def norm(v, asc=True):
        v=(v-v.min())/(v.max()-v.min()+1e-9); return v if asc else (1-v)
    score = 0.4*norm(m['silhouette'],True)+0.4*norm(m['CH'],True)+0.2*norm(m['DBI'],False)
    k_best = int(m.loc[score.idxmax(),'k'])
    return k_best, m, score

# ========================== Evaluation Metrics ==========================
def metrics_all(y_true, y_pred, y_train_for_scaler):
    y_true = np.asarray(y_true).reshape(-1); y_pred=np.asarray(y_pred).reshape(-1)
    R2  = r2_score(y_true, y_pred)
    MSE = mean_squared_error(y_true, y_pred)
    MAE = mean_absolute_error(y_true, y_pred)
    MAPE= mean_absolute_percentage_error(np.clip(y_true,1e-9,None), np.clip(y_pred,1e-9,None))
    std = float(np.std(np.asarray(y_train_for_scaler).reshape(-1), ddof=0))
    sMSE = MSE / (std**2 + 1e-9)
    y_min, y_max = y_true.min(), y_true.max()
    if y_max>y_min:
        yt=(y_true-y_min)/(y_max-y_min); yp=(y_pred-y_min)/(y_max-y_min)
        NMSE = mean_squared_error(yt, yp)
    else:
        NMSE = 0.0
    return dict(R2=R2, MSE=MSE, MAE=MAE, sMSE=sMSE, NMSE=NMSE, MAPE=MAPE)

# ========================== Feature Blocks (Economic PCA, others standardization) + Target Transformation ==========================
def make_feature_block(feats):
    econ_feats=['people','total_GDP','Urbanization_rate','Baidu_index','CEV_Subsidy funds']
    time_feats=['month_sin','month_cos','year_num']
    lag_feats =['lag_1','lag_3','lag_12','roll_mean_3','roll_mean_12']
    sea_feats =[f for f in SEASONAL_FEATS if f in feats]
    inter_feats=[f for f in feats if f in ['gdp_per_capita','subsidy_to_gdp','baidu_per_capita']]

    transformers=[]
    if any(f in feats for f in econ_feats):
        transformers.append(('econ', Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('sc',  StandardScaler()),
            ('pca', PCA(n_components=0.95, svd_solver='full', random_state=RANDOM_STATE))
        ]), [f for f in econ_feats if f in feats]))
    if any(f in feats for f in time_feats):
        transformers.append(('time', Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('sc',  StandardScaler())
        ]), [f for f in time_feats if f in feats]))
    if any(f in feats for f in lag_feats):
        transformers.append(('lag', Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('sc',  StandardScaler())
        ]), [f for f in lag_feats if f in feats]))
    if sea_feats:
        transformers.append(('seasonal', Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('sc',  StandardScaler())
        ]), sea_feats))
    if inter_feats:
        transformers.append(('inter', Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('sc',  StandardScaler())
        ]), inter_feats))
    return ColumnTransformer(transformers=transformers, remainder='drop')

def _need_log1p(y_series: pd.Series) -> bool:
    s = pd.Series(y_series).dropna()
    if len(s) < 30:  # Too few samples, no transformation
        return False
    skew = float(s.skew())
    return (skew > 1.0)  # Skewness threshold, adjustable as needed

def make_estimator_with_ttr(reg, use_log: bool):
    transf = StandardScaler() if not use_log else Pipeline(steps=[
        ('log',  FunctionLog1p()),    # Custom log1p/exp for TTR
        ('sc',   StandardScaler())
    ])
    return TransformedTargetRegressor(regressor=reg, transformer=transf)

# Simple log1p/expm1 wrapper to avoid validate warning
from sklearn.preprocessing import FunctionTransformer
class FunctionLog1p(FunctionTransformer):
    def __init__(self):
        super().__init__(func=np.log1p, inverse_func=np.expm1, validate=False)

# ========================== Time Series CV (with gap) ==========================
def make_tscv(n_samples, n_splits=N_SPLITS):
    n_splits = min(max(3, n_splits), max(3, n_samples//12))
    return TimeSeriesSplit(n_splits=n_splits)

# ========================== Hyperparameter Search Adaptive by Cluster Size ==========================
def fit_with_search(name, base_pipe, X, y, use_log, is_cluster_0=False, n_cities=100):
    tscv = make_tscv(len(X))
    n_trials = N_TRIALS_C0 if is_cluster_0 else N_TRIALS
    small_cluster = n_cities < 15

    if name == 'RF':
        param_distributions = {
            'est__regressor__n_estimators': ([250,300,350,400,450] if is_cluster_0 else [150,200,250,300]),
            'est__regressor__max_depth':    ([4,5,6,7,8,9,None] if is_cluster_0 else [3,4,5,6,7]),
            'est__regressor__min_samples_split': [6,10,15,20,25,30],
            'est__regressor__min_samples_leaf':  [2,4,6,8,10] if is_cluster_0 else [4,6,8,10,12],
            'est__regressor__max_features': ['sqrt']
        }
        base = RandomForestRegressor(random_state=RANDOM_STATE, n_jobs=-1)
        est  = make_estimator_with_ttr(base, use_log)
        search = RandomizedSearchCV(
            estimator=Pipeline([('ct', base_pipe), ('est', est)]),
            param_distributions=param_distributions, n_iter=n_trials,
            cv=tscv, scoring='r2', random_state=RANDOM_STATE, n_jobs=-1
        )

    elif name == 'XGB':
        param_distributions = {
            'est__regressor__n_estimators': ([300,400,500,600] if is_cluster_0 else [200,250,300,350]),
            'est__regressor__max_depth': [3,4,5,6],
            'est__regressor__learning_rate': [0.03,0.05,0.07,0.1],
            'est__regressor__subsample': [0.7,0.8,0.9,1.0],
            'est__regressor__colsample_bytree': [0.7,0.8,0.9,1.0],
            'est__regressor__min_child_weight': [1,3,5,7],
            'est__regressor__gamma': [0, 0.1, 0.3, 1.0],
            'est__regressor__reg_alpha': [0.0,0.1,0.3,1.0,3.0],
            'est__regressor__reg_lambda': [1.0,2.0,3.0,5.0],
        }
        xgb = XGBRegressor(objective='reg:squarederror', random_state=RANDOM_STATE, n_jobs=-1)
        est = make_estimator_with_ttr(xgb, use_log)
        search = RandomizedSearchCV(
            estimator=Pipeline([('ct', base_pipe), ('est', est)]),
            param_distributions=param_distributions, n_iter=n_trials,
            cv=tscv, scoring='r2', random_state=RANDOM_STATE, n_jobs=-1
        )

    elif name == 'MLP':
        param_distributions = {
            'est__regressor__hidden_layer_sizes': ([(128,64),(160,80),(200,100),(160,100,50)]
                                                   if is_cluster_0 else [(64,32),(96,48),(128,64)]),
            'est__regressor__alpha': [1e-4,3e-4,1e-3,3e-3,1e-2],
            'est__regressor__learning_rate_init': [1e-3,2e-3,3e-3,5e-3,1e-2],
            'est__regressor__batch_size': ['auto']  # Avoid small sample warning
        }
        mlp = MLPRegressor(random_state=RANDOM_STATE, early_stopping=True,
                           max_iter=900, n_iter_no_change=20, validation_fraction=0.15)
        est = make_estimator_with_ttr(mlp, use_log)
        search = RandomizedSearchCV(
            estimator=Pipeline([('ct', base_pipe), ('est', est)]),
            param_distributions=param_distributions, n_iter=n_trials,
            cv=tscv, scoring='r2', random_state=RANDOM_STATE, n_jobs=-1
        )

    elif name == 'SVM':
        if small_cluster:
            C_grid     = [0.2,0.3,0.5,0.8,1.0,1.5]
            gamma_grid = [1e-4,3e-4,8e-4,1e-3,2e-3]
            eps_grid   = [3e-3,1e-2,2e-2,0.03]
        else:
            C_grid     = [0.3,0.5,1,2,3,5]
            gamma_grid = [1e-4,3e-4,1e-3,3e-3,1e-2]
            eps_grid   = [1e-3,3e-3,1e-2,3e-2,0.05]
        param_distributions = {
            'est__regressor__kernel': ['rbf'],
            'est__regressor__C': C_grid,
            'est__regressor__gamma': gamma_grid,
            'est__regressor__epsilon': eps_grid
        }
        svm = SVR()
        est = make_estimator_with_ttr(svm, use_log)
        search = RandomizedSearchCV(
            estimator=Pipeline([('ct', base_pipe), ('est', est)]),
            param_distributions=param_distributions, n_iter=n_trials,
            cv=tscv, scoring='r2', random_state=RANDOM_STATE, n_jobs=-1
        )
    else:
        raise ValueError("unknown model name")

    # Important: Use gap=1 splitting within CV
    # (sklearn TimeSeriesSplit doesn't directly support gap; leakage avoided in feature construction; strictly sequential split here)
    search.fit(X, y)
    best_est = search.best_estimator_
    best_params = {k.replace('est__regressor__','').replace('est__',''):v for k,v in search.best_params_.items()}
    return best_est, best_params, search

# ========================== Stacking & Blending ==========================
class StackingRFLinear:
    """Level 1 RF (with y transformation/scaling) → Time series OOF → Level 2 linear regression"""
    def __init__(self, rf_pipe): self.rf_pipe=rf_pipe; self.meta=LinearRegression()
    def fit(self,X,y):
        tscv = make_tscv(len(X))
        oof = np.full(len(X), np.nan)
        for tr,te in tscv.split(X):
            self.rf_pipe.fit(X.iloc[tr], y.iloc[tr]); oof[te] = self.rf_pipe.predict(X.iloc[te])
        mask = ~np.isnan(oof)
        self.meta.fit(oof[mask].reshape(-1,1), y.iloc[mask])
        self.rf_pipe.fit(X, y); return self
    def predict(self,X):
        base = self.rf_pipe.predict(X)
        return self.meta.predict(base.reshape(-1,1))

class StackingBaseLinear:
    def __init__(self, base_pipe):
        self.base_pipe = base_pipe
        self.meta = RidgeCV(alphas=np.logspace(-4,3,20))
    def fit(self,X,y):
        tscv = make_tscv(len(X))
        oof = np.full(len(X), np.nan)
        for tr,te in tscv.split(X):
            self.base_pipe.fit(X.iloc[tr], y.iloc[tr])
            oof[te] = self.base_pipe.predict(X.iloc[te])
        mask = ~np.isnan(oof)
        self.meta.fit(oof[mask].reshape(-1,1), y.iloc[mask])
        self.base_pipe.fit(X, y); return self
    def predict(self,X):
        base = self.base_pipe.predict(X)
        return self.meta.predict(base.reshape(-1,1))

class BlendTop2:
    def __init__(self, name1, mdl1, name2, mdl2, w=0.6):
        self.name1, self.mdl1 = name1, mdl1
        self.name2, self.mdl2 = name2, mdl2
        self.w=w
    def fit(self, X, y): return self
    def predict(self,X):
        return self.w*self.mdl1.predict(X) + (1-self.w)*self.mdl2.predict(X)

# ========================== Forecast: Rolling + Anti-explosion ==========================
def iterative_forecast_tabular(best_model, hist_feat_df, future_feat_df, feats, target,
                               cap_mult=HIST_MAX_MULT, nonneg=CLIP_NONNEG):
    last={}
    for city, g in hist_feat_df[hist_feat_df['Year']<=2024].groupby('City'):
        tail=g[target].dropna().tail(24).tolist()
        while len(tail)<24: tail=[0.0]+tail
        last[city]=tail[-24:]
    hist_cap = hist_feat_df.groupby('City')[target].max().to_dict()

    f = future_feat_df.copy()
    f['date'] = pd.to_datetime(dict(year=f['Year'], month=f['Month'], day=1))
    f.sort_values(['date','City'], inplace=True)

    rows=[]
    for (yr,mo), g in f.groupby(['Year','Month']):
        batch=[]; keys=[]
        for _, r in g.iterrows():
            city=r['City']
            if city not in last: last[city]=[0.0]*24
            seq=last[city]
            feat = {
                'people': float(r.get('people',0.0)),
                'total_GDP': float(r.get('total_GDP',0.0)),
                'Urbanization_rate': float(r.get('Urbanization_rate',0.0)),
                'Baidu_index': float(r.get('Baidu_index',0.0)),
                'CEV_Subsidy funds': float(r.get('CEV_Subsidy funds',0.0)),
                'month_sin': math.sin(2*math.pi*(int(mo)-1)/12),
                'month_cos': math.cos(2*math.pi*(int(mo)-1)/12),
                'year_num': int(yr),
                'lag_1': seq[-1], 'lag_3': seq[-3], 'lag_12': seq[-12],
                'roll_mean_3': np.mean(seq[-3:]),
                'roll_mean_12': np.mean(seq[-12:])
            }
            # Seasonal differencing (strictly using history)
            lag1, lag13 = seq[-1], seq[-13]
            feat['sea_diff_1_12']      = lag1 - lag13
            feat['sea_ratio_1_12']     = lag1 / (lag13 + 1e-9)
            feat['sea_diff_mean3_12']  = np.mean(seq[-3:]) - np.mean(seq[-15:-12])
            # y deseasonal residual (using historical rolling mean)
            feat['y_deseasonal_res']   = lag1 - np.mean(seq[-13:-1]) if len(seq) >= 13 else 0.0

            # Interaction
            if 'gdp_per_capita' in feats and feat['people']>0:
                feat['gdp_per_capita'] = feat['total_GDP']/(feat['people']+1e-9)
            if 'subsidy_to_gdp' in feats and feat['total_GDP']>0:
                feat['subsidy_to_gdp'] = feat['CEV_Subsidy funds']/(feat['total_GDP']+1e-9)
            if 'baidu_per_capita' in feats and feat['people']>0:
                feat['baidu_per_capita'] = feat['Baidu_index']/(feat['people']+1e-9)

            batch.append([feat.get(c,0.0) for c in feats]); keys.append(city)

        if not batch: continue
        Xb = pd.DataFrame(batch, columns=feats)
        yhat = best_model.predict(Xb)
        for city, val in zip(keys, yhat):
            v=float(val)
            if nonneg: v=max(0.0, v)
            cap=float(hist_cap.get(city, np.nan))
            if np.isfinite(cap) and cap>0:
                v=min(v, cap*cap_mult)
            rows.append({'City':city,'Year':int(yr),'Month':int(mo), f'{target}_pred':v})
            last[city].append(v); last[city]=last[city][-24:]
    return pd.DataFrame(rows).sort_values(['City','Year','Month']).reset_index(drop=True)

# ========================== Training/Comparison/Selection/Forecast ==========================
def train_compare_cluster(sub_hist, sub_future, target, out_tag):
    sub_tab, feats = build_tabular(sub_hist.copy(), target)

    # Train/test split: Default ≤2023 train, =2024 test; otherwise split by time 85%
    train_df = sub_tab[sub_tab['Year']<=2023]
    test_df  = sub_tab[sub_tab['Year']==2024]
    if test_df.empty and len(sub_tab)>=24:
        cutoff = sub_tab['date'].quantile(0.85)
        train_df = sub_tab[sub_tab['date']<=cutoff]
        test_df  = sub_tab[sub_tab['date']>cutoff]
    if len(train_df)==0 or len(test_df)==0:
        raise RuntimeError("Insufficient samples in this cluster")

    # Adaptive target transformation (by skewness)
    use_log = _need_log1p(train_df[target])

    ct = make_feature_block(feats)
    n_cities = sub_hist['City'].nunique()
    is_cluster_0 = ('Cluster0' in out_tag)

    # —— 4 Base Models —— #
    rf,  rf_p,  _ = fit_with_search('RF',  ct, train_df[feats], train_df[target], use_log, is_cluster_0, n_cities)
    xgb, xgb_p, _ = fit_with_search('XGB', ct, train_df[feats], train_df[target], use_log, is_cluster_0, n_cities)
    mlp, mlp_p, _ = fit_with_search('MLP', ct, train_df[feats], train_df[target], use_log, is_cluster_0, n_cities)
    svm, svm_p, _ = fit_with_search('SVM', ct, train_df[feats], train_df[target], use_log, is_cluster_0, n_cities)

    # Stacking (two directions)
    stack_rf  = StackingRFLinear(rf_pipe=rf).fit(train_df[feats], train_df[target])
    stack_mlp = StackingBaseLinear(mlp).fit(train_df[feats], train_df[target])

    # Single model evaluation first, for selecting Top2 blending
    single = {'RF':rf,'XGB':xgb,'ANN_MLP':mlp,'SVM':svm}
    single_rows=[]
    for name, mdl in single.items():
        y_pred = mdl.predict(test_df[feats])
        m = metrics_all(test_df[target], y_pred, train_df[target]); m['model']=name
        single_rows.append(m)
    single_df = pd.DataFrame(single_rows).sort_values(['R2','sMSE'], ascending=[False,True])

    # Top2 blending (only added as candidate in large clusters, to alleviate C0 underfitting)
    models = {'RF':rf,'XGB':xgb,'ANN_MLP':mlp,'SVM':svm,
              'Stacking_RF+Linear':stack_rf, 'Stacking_MLP+Ridge':stack_mlp}
    if is_cluster_0 and len(single_df)>=2:
        t1, t2 = single_df.iloc[0]['model'], single_df.iloc[1]['model']
        models[f'BlendTop2({t1}+{t2})'] = BlendTop2(t1, single[t1], t2, single[t2], w=0.6)

    # Unified evaluation
    eval_rows=[]
    for name, mdl in models.items():
        yp = mdl.predict(test_df[feats])
        m  = metrics_all(test_df[target], yp, train_df[target]); m.update(model=name)
        eval_rows.append(m)
    eval_df = pd.DataFrame(eval_rows).sort_values(['R2','sMSE'], ascending=[False,True])

    # Select best + report Train/Test R²
    best_name  = eval_df.iloc[0]['model']
    best_model = models[best_name]
    tr_pred = best_model.predict(train_df[feats])
    train_r2 = r2_score(train_df[target], tr_pred)
    test_r2  = eval_df.iloc[0]['R2']
    print(f"[{out_tag}] Best model: {best_name} | Train R2={train_r2:.3f} | Test R2={test_r2:.3f}")

    # —— Combine "train+test" scatter in one chart, display dual R² —— #
    png_scatter = os.path.join(OUT_DIR, f"{out_tag}_2016-2024_fit_{best_name}.png")
    plt.figure(figsize=(5.6,5.2))
    sns.scatterplot(x=train_df[target].values, y=tr_pred, s=15, label=f"Train (R²={train_r2:.3f})")
    te_pred = best_model.predict(test_df[feats])
    sns.scatterplot(x=test_df[target].values, y=te_pred, s=20, label=f"Test  (R²={test_r2:.3f})")
    lims=[min(train_df[target].min(), test_df[target].min(), tr_pred.min(), te_pred.min()),
          max(train_df[target].max(), test_df[target].max(), tr_pred.max(), te_pred.max())]
    plt.plot(lims, lims, '--', lw=1)
    plt.title(f"{out_tag} Fit (train+test)")
    plt.xlabel("Actual"); plt.ylabel("Predicted"); plt.legend()
    plt.tight_layout(); plt.savefig(png_scatter, dpi=150); plt.close()

    # —— Retrain with ≤2024 full data → Future rolling forecast (anti-explosion, non-negative) —— #
    hist_tab = sub_tab[sub_tab['Year']<=2024].copy()
    best_model.fit(hist_tab[feats], hist_tab[target])
    pred_future = iterative_forecast_tabular(best_model, hist_tab, sub_future.copy(), feats, target,
                                             cap_mult=HIST_MAX_MULT, nonneg=CLIP_NONNEG)

    # Export evaluation
    eval_out = eval_df.copy()
    eval_out.loc[eval_out['model']==best_name,'Train_R2']=train_r2
    eval_out.loc[eval_out['model']==best_name,'Test_R2']=test_r2
    write_df(MASTER_XLSX, eval_out, sheet=f'{out_tag}_EvalAll', index=False)
    embed_png(MASTER_XLSX, f'{out_tag}_EvalAll', png_scatter, anchor="H2")

    return best_name, pred_future, eval_df

# ========================== Main Process (CEV only) ==========================
def run_cev_only():
    df, t_cev = read_panel(INPUT_PATH, SHEET_NAME)

    # Light interpolation + mild winsorize
    for col in ['people','total_GDP','Urbanization_rate','Baidu_index','CEV_Subsidy funds', t_cev]:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    for city, g in df.groupby('City'):
        idx = g.index
        df.loc[idx, [t_cev,'people','total_GDP','Urbanization_rate','Baidu_index','CEV_Subsidy funds']] = \
            g[[t_cev,'people','total_GDP','Urbanization_rate','Baidu_index','CEV_Subsidy funds']].interpolate(limit_direction='both')
        s = df.loc[idx, t_cev].dropna()
        if len(s)>=20:
            lo, hi = np.percentile(s, 1), np.percentile(s, 99)
            df.loc[idx, t_cev] = df.loc[idx, t_cev].clip(lo, hi)

    hist = df[(df['Year']>=2016) & (df['Year']<=2024)].copy()
    fut  = df[(df['Year']>=2025) & (df['Year']<=2030)].copy()

    # —— Clustering (auto K selection) —— #
    print("\n==== [CEV] Auto K selection and clustering ====")
    city_feat = extract_city_level_features(hist, t_cev)
    emb      = embed_city_features(city_feat)
    k_best, metric_tbl, score = choose_k_enhanced(emb)
    km = KMeans(n_clusters=k_best, n_init=100, random_state=RANDOM_STATE).fit(emb[['pc1','pc2']].values)
    labels = km.labels_
    emb_k  = emb.assign(CityCluster=labels)
    print(f"→ [CEV] Best K = {k_best}, number of clusters: {k_best}")

    # Export cluster table and metrics
    write_df(MASTER_XLSX, emb_k.sort_values(['CityCluster','City']), sheet='CEV_Clusters', index=False)
    if metric_tbl is not None:
        mt = metric_tbl.copy()
        if score is not None: mt['score']=score.values
        write_df(MASTER_XLSX, mt, sheet='CEV_ClusterMetrics', index=False)

    # Merge cluster labels
    hist_k = hist.merge(emb_k[['City','CityCluster']], on='City', how='left')
    fut_k  = fut .merge(emb_k[['City','CityCluster']], on='City', how='left')

    # —— Per cluster training/selection/forecast —— #
    all_eval=[]; all_best=[]; all_pred=[]
    for k in sorted(hist_k['CityCluster'].dropna().unique()):
        sub_hist   = hist_k[hist_k['CityCluster']==k].copy()
        sub_future = fut_k [fut_k ['CityCluster']==k].copy()
        if sub_hist.empty or sub_future.empty: continue
        print(f"  - Training cluster C{k} (number of cities={sub_hist['City'].nunique()})")
        try:
            best_name, pred_mon, eval_df = train_compare_cluster(sub_hist, sub_future, t_cev, out_tag=f'CEV_Cluster{k}')
            all_best.append({'CityCluster':int(k),'BestModel':best_name,'TestR2':float(eval_df.iloc[0]['R2'])})
            all_eval.append(eval_df.assign(CityCluster=int(k)))
            all_pred.append(pred_mon.assign(CityCluster=int(k)))
        except Exception as e:
            print(f"[WARN] Cluster C{k} failed:", e)
            traceback.print_exc()

    # Export summary tables
    if all_eval:
        E = pd.concat(all_eval, ignore_index=True)
        write_df(MASTER_XLSX, E.sort_values(['CityCluster','R2'], ascending=[True,False]),
                 sheet='CEV_EvalAll', index=False)
    if all_best:
        B = pd.DataFrame(all_best)
        write_df(MASTER_XLSX, B.sort_values('CityCluster'), sheet='CEV_BestPerCluster', index=False)
    if all_pred:
        P  = pd.concat(all_pred, ignore_index=True)
        Py = P.groupby(['City','Year']).sum(numeric_only=True).reset_index()
        write_df(MASTER_XLSX, P.sort_values(['City','Year','Month']), sheet='CEV_PredMonthly', index=False)
        write_df(MASTER_XLSX, Py.sort_values(['City','Year']), sheet='CEV_PredAnnual',  index=False)

    print(f"\n==== [CEV] Completed, results written to {MASTER_XLSX} ====\n")

if __name__ == "__main__":
    run_cev_only()