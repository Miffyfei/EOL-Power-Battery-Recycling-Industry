# -*- coding: utf-8 -*-
"""
PEV-best: Clustering (Auto K selection) → Cluster-wise training/evaluation/prediction
Key changes:
- C0/C1/C2 use "locked models and hyperparameters" to reproduce and stabilize previous best metrics (avoid fluctuations from re-searching)
- Added seasonal difference feature sdiff_12 = y_t - y_{t-12}
- Still output R2/MSE/MAE/sMSE/NMSE/MAPE + charts + Excel, multi-model evaluation but only search on non-locked clusters
"""

import os, json, math, warnings, traceback
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

from sklearn.model_selection import TimeSeriesSplit, RandomizedSearchCV
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer, TransformedTargetRegressor
from sklearn.metrics import (r2_score, mean_squared_error, mean_absolute_error,
                             silhouette_score, silhouette_samples,
                             davies_bouldin_score, calinski_harabasz_score,
                             mean_absolute_percentage_error)
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans

from sklearn.ensemble import RandomForestRegressor
from sklearn.neural_network import MLPRegressor
from sklearn.linear_model import RidgeCV
from sklearn.svm import SVR
from xgboost import XGBRegressor

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

# ========================== Paths and Global Configuration ==========================
INPUT_PATH = "./input data/2016-2030_PEV and CEV_month+data.xlsx"
SHEET_NAME = "Sheet1"
OUT_DIR = "./output data/results_electric vehicle"
os.makedirs(OUT_DIR, exist_ok=True)
MASTER_XLSX = os.path.join(OUT_DIR, "PEV_ClusterBest_Forecast.xlsx")

RANDOM_STATE = 42
np.random.seed(RANDOM_STATE)

# Search intensity (only for non-locked clusters)
N_TRIALS = 120
N_TRIALS_C0 = 150
N_SPLITS = 5
K_RANGE = range(3, 15)

# Target column candidates (PEV only)
TARGET_PEV_CANDS = ['PEV_number', 'Commercial_on_insurance', 'EV_Commercial_Sales', 'PEV']

# Base feature mapping (fault-tolerant)
BASE_FEATURE_CANDS = {
    'people': ['people', 'Population', 'population', 'People'],
    'total_GDP': ['total_GDP', 'GDP', 'gdp', 'Total_GDP'],
    'Urbanization_rate': ['Urbanization_rate', 'urbanization_rate', 'Urban_rate'],
    'Baidu_index': ['Baidu_index', 'baidu_index', 'Baidu'],
    'PEV_Subsidy funds': ['PEV_Subsidy funds', 'Subsidy', 'PEV_Subsidy', 'subsidy']
}

# Tabular base features
TAB_FEATS = [
    'people','total_GDP','Urbanization_rate','Baidu_index','PEV_Subsidy funds',
    'month_sin','month_cos','year_num',
    'lag_1','lag_3','lag_12','roll_mean_3','roll_mean_12',
    # Added: seasonal difference
    'sdiff_12'
]

# ---------------- Locked model parameters (from your provided "best configuration") ----------------
LOCKED = {
    0: {  # Cluster 0
        "type": "BlendTop2",
        "blend_weight": 0.6,
        "ANN_MLP_params": {"learning_rate_init": 0.003, "hidden_layer_sizes": (150, 100, 50), "batch_size": 32, "alpha": 0.0005},
        "SVM_params":     {"kernel": "rbf", "gamma": 1e-4, "epsilon": 0.05, "C": 3}
    },
    1: {  # Cluster 1
        "type": "MLP",
        "MLP_params": {"learning_rate_init": 0.001, "hidden_layer_sizes": (150, 75), "batch_size": 64, "alpha": 0.0001}
    },
    2: {  # Cluster 2
        "type": "MLP",
        "MLP_params": {"learning_rate_init": 0.01, "hidden_layer_sizes": (150, 75), "batch_size": 32, "alpha": 0.01}
    }
}
# --------------------------------------------------------------------

# =============== Excel I/O ===============
def ensure_book(path):
    if not os.path.exists(path):
        wb = Workbook(); wb.save(path)

def write_df(path, df: pd.DataFrame, sheet: str, index=False):
    ensure_book(path)
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
        df.to_excel(w, sheet_name=sheet, index=index)

def embed_png(path, sheet, img_path, anchor="A1"):
    if not os.path.exists(img_path): return
    ensure_book(path)
    wb = load_workbook(path)
    if sheet not in wb.sheetnames: wb.create_sheet(sheet)
    ws = wb[sheet]
    try:
        img = XLImage(img_path); img.anchor = anchor
        ws.add_image(img)
    except Exception as e:
        print("Image embedding failed:", e)
    wb.save(path)

# =============== Data Reading ===============
def pick_col(df, cands):
    ok = [(c, df[c].notna().sum()) for c in cands if c in df.columns]
    if not ok: return None
    ok.sort(key=lambda x: x[1], reverse=True)
    return ok[0][0]

def read_panel(path, sheet):
    df = pd.read_excel(path, sheet_name=sheet)
    df.columns = df.columns.map(lambda x: str(x).strip())
    for base in ['City','Year','Month']:
        if base not in df.columns:
            hits = [c for c in df.columns if c.lower()==base.lower()]
            if hits: df.rename(columns={hits[0]:base}, inplace=True)
    assert {'City','Year','Month'}.issubset(df.columns), "Require City/Year/Month columns"

    for std, cands in BASE_FEATURE_CANDS.items():
        if std not in df.columns:
            hit = [c for c in cands if c in df.columns]
            if hit: df.rename(columns={hit[0]:std}, inplace=True)
        if std not in df.columns: df[std] = np.nan

    t_pev = pick_col(df, TARGET_PEV_CANDS)
    if t_pev is None:
        raise ValueError("PEV target column not found")

    for c in ['people','total_GDP','Urbanization_rate','Baidu_index','PEV_Subsidy funds', t_pev]:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    df['date'] = pd.to_datetime(dict(year=df['Year'].astype(int), month=df['Month'].astype(int), day=1))
    df = df.sort_values(['City','date']).reset_index(drop=True)
    return df, t_pev

# =============== Feature Engineering (including seasonal difference sdiff_12) ===============
def add_time_cols(df):
    df = df.copy()
    m = df['Month'].astype(int)
    df['month_sin'] = np.sin(2*np.pi*(m-1)/12)
    df['month_cos'] = np.cos(2*np.pi*(m-1)/12)
    df['year_num']  = df['Year'].astype(int)
    return df

def add_lag_roll(df, y, lags=(1,3,12), wins=(3,12)):
    df = df.sort_values(['City','date']).copy()
    for city, g in df.groupby('City'):
        idx = g.index
        for L in lags:
            df.loc[idx, f'lag_{L}'] = g[y].shift(L)
        for w in wins:
            df.loc[idx, f'roll_mean_{w}'] = g[y].shift(1).rolling(w).mean()
        # Seasonal difference
        df.loc[idx, 'sdiff_12'] = g[y] - g[y].shift(12)
    return df

def build_tabular(df, y):
    df = add_time_cols(df)
    df = add_lag_roll(df, y)

    # Interaction features
    if 'total_GDP' in df.columns and 'people' in df.columns:
        df['gdp_per_capita'] = df['total_GDP'] / (df['people'] + 1e-9)
    if 'PEV_Subsidy funds' in df.columns and 'total_GDP' in df.columns:
        df['subsidy_to_gdp'] = df['PEV_Subsidy funds'] / (df['total_GDP'] + 1e-9)
    if 'Baidu_index' in df.columns and 'people' in df.columns:
        df['baidu_per_capita'] = df['Baidu_index'] / (df['people'] + 1e-9)

    for c in TAB_FEATS:
        if c not in df.columns: df[c] = np.nan
        df[c] = pd.to_numeric(df[c], errors='coerce')

    # sdiff_12 allows NaN (first 12 periods), models will use median imputation; lag features still ensure non-null
    df = df.dropna(subset=['lag_1','lag_3','lag_12'])
    extra = ['gdp_per_capita','subsidy_to_gdp','baidu_per_capita']
    feats = TAB_FEATS + [f for f in extra if f in df.columns]
    return df, feats

# =============== Correlation (for interpretation) ===============
def corr_abs_to_target(df, target):
    cols = list({*TAB_FEATS, target})
    extended_cols = [c for c in df.columns if (c in cols) or ('per_capita' in c) or ('to_gdp' in c)]
    sub = add_time_cols(df)[[c for c in extended_cols if c in df.columns]].copy()
    C = sub.corr(numeric_only=True)
    if target not in C.columns: return pd.Series(dtype=float)
    return C[target].drop(index=target).abs().sort_values(ascending=False)

# =============== Clustering (Auto K selection + Visualization) ===============
def extract_city_level_features(df, target):
    d = df.copy()
    d['Urban_population'] = d['people'] * d['Urbanization_rate']
    bases = ['total_GDP','Urban_population','Baidu_index','people','Urbanization_rate']
    rows=[]
    for city, g in d.groupby('City'):
        g = g.sort_values('date')
        r = {'City':city}
        g_hist = g[g['Year']<=2024]
        for b in bases: r[f'lv_{b}'] = g_hist[g_hist['Year']==2024][b].mean()
        g_recent = g_hist[g_hist['Year']>=2022]
        for b in bases:
            s = g_recent[b].dropna()
            if len(s)>=2:
                first,last = s.iloc[0], s.iloc[-1]
                n = max(1,len(s)-1)
                cagr = ((last+1e-9)/(first+1e-9))**(1/n)-1
                r[f'cagr_{b}']=cagr; r[f'var_{b}']=s.var()
            else:
                r[f'cagr_{b}']=0.0; r[f'var_{b}']=0.0
        rows.append(r)
    return pd.DataFrame(rows).fillna(0.0)

def embed_city_features(city_feat):
    scaler = StandardScaler()
    X = scaler.fit_transform(city_feat.drop(columns=['City']))
    pca = PCA(n_components=2, random_state=RANDOM_STATE)
    Xp = pca.fit_transform(X)
    return city_feat[['City']].assign(pc1=Xp[:,0], pc2=Xp[:,1])

def choose_k_enhanced(emb, k_range=K_RANGE):
    X = emb[['pc1','pc2']].values
    rows=[]
    for k in k_range:
        km = KMeans(n_clusters=k, n_init=100, random_state=RANDOM_STATE).fit(X)
        labs = km.labels_
        if any((labs==i).sum()<3 for i in range(k)):  # At least 3 cities per cluster
            continue
        sil = silhouette_score(X, labs)
        dbi = davies_bouldin_score(X, labs)
        ch  = calinski_harabasz_score(X, labs)
        wcss = km.inertia_
        rows.append({'k':k,'silhouette':sil,'DBI':dbi,'CH':ch,'WCSS':wcss})
    m = pd.DataFrame(rows)
    if m.empty:  # Fallback
        return 3, None, None
    def norm(v, asc=True):
        v=(v-v.min())/(v.max()-v.min()+1e-9)
        return v if asc else (1-v)
    score = (norm(m['silhouette'],True)*0.4 + norm(m['CH'],True)*0.4 + norm(m['DBI'],False)*0.2)
    k_best = int(m.loc[score.idxmax(),'k'])
    return k_best, m, score

def plot_cluster_scatter(emb, labels, path_png, title="Clusters (PCA-2D)"):
    plt.figure(figsize=(7.2,5.2))
    for k in sorted(np.unique(labels)):
        g = emb[labels==k]
        plt.scatter(g['pc1'], g['pc2'], s=22, label=f'C{k}')
    plt.legend(); plt.title(title); plt.tight_layout(); plt.savefig(path_png, dpi=150); plt.close()

def plot_silhouette_bars(emb, labels, path_png):
    X = emb[['pc1','pc2']].values
    s = silhouette_samples(X, labels); avg_s = s.mean()
    plt.figure(figsize=(7.2,5.2)); y_lower=10
    for k in sorted(np.unique(labels)):
        s_k = s[labels==k]; s_k.sort()
        size=len(s_k); y_upper=y_lower+size
        color = plt.cm.tab10(k % 10)
        plt.fill_betweenx(np.arange(y_lower, y_upper), 0, s_k, facecolor=color, alpha=0.8)
        plt.text(-0.05, y_lower+0.5*size, f"C{k}")
        y_lower = y_upper + 10
    plt.axvline(avg_s, color='red', linestyle='--', label=f"avg={avg_s:.3f}")
    plt.xlabel("Silhouette coefficient"); plt.ylabel("Samples"); plt.legend()
    plt.tight_layout(); plt.savefig(path_png, dpi=150); plt.close()

# =============== Metrics (including MAPE) ===============
def metrics_all(y_true, y_pred, y_train_for_scaler):
    y_true = np.asarray(y_true).reshape(-1); y_pred=np.asarray(y_pred).reshape(-1)
    R2  = r2_score(y_true, y_pred)
    MSE = mean_squared_error(y_true, y_pred)
    MAE = mean_absolute_error(y_true, y_pred)
    MAPE = mean_absolute_percentage_error(y_true, y_pred)
    std = float(np.std(np.asarray(y_train_for_scaler).reshape(-1), ddof=0))
    sMSE = MSE / (std**2 + 1e-9)
    y_min, y_max = y_true.min(), y_true.max()
    if y_max>y_min:
        yt = (y_true - y_min)/(y_max-y_min)
        yp = (y_pred - y_min)/(y_max-y_min)
        NMSE = mean_squared_error(yt, yp)
    else:
        NMSE = 0.0
    return dict(R2=R2, MSE=MSE, MAE=MAE, sMSE=sMSE, NMSE=NMSE, MAPE=MAPE)

# =============== Feature Blocks (PCA only on economic block) ===============
def make_feature_block(feats):
    econ_feats = ['people','total_GDP','Urbanization_rate','Baidu_index','PEV_Subsidy funds']
    time_feats = ['month_sin','month_cos','year_num']
    lag_feats  = ['lag_1','lag_3','lag_12','roll_mean_3','roll_mean_12','sdiff_12']
    inter_feats = [f for f in feats if 'per_capita' in f or 'to_gdp' in f]

    econ = [f for f in econ_feats if f in feats]
    tim  = [f for f in time_feats if f in feats]
    lag  = [f for f in lag_feats if f in feats]
    inter = [f for f in inter_feats if f in feats]

    transformers=[]
    if econ:
        transformers.append(('econ', Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('sc', StandardScaler()),
            ('pca', PCA(n_components=0.95, svd_solver='full', random_state=RANDOM_STATE))
        ]), econ))
    if tim:
        transformers.append(('time', Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('sc', StandardScaler()),
        ]), tim))
    if lag:
        transformers.append(('lag', Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('sc', StandardScaler()),
        ]), lag))
    if inter:
        transformers.append(('interaction', Pipeline([
            ('imp', SimpleImputer(strategy='median')),
            ('sc', StandardScaler()),
        ]), inter))
    return ColumnTransformer(transformers=transformers, remainder='drop')

def wrap_y(estimator):
    return TransformedTargetRegressor(regressor=estimator, transformer=StandardScaler())

# =============== Search (only for non-locked clusters, avoid overfitting fluctuations) ===============
def make_tscv(n_samples, n_splits=N_SPLITS):
    return TimeSeriesSplit(n_splits=min(max(2, n_splits), max(2, n_samples // 12)))

def fit_with_search(name, base_pipe, X, y, is_cluster_0=False):
    tscv = make_tscv(len(X))
    n_trials = N_TRIALS_C0 if is_cluster_0 else N_TRIALS

    if name == 'MLP':
        param_distributions = {
            'est__regressor__hidden_layer_sizes': ([(128,64),(150,75),(200,100),(150,100,50)] if is_cluster_0 else [(64,32),(100,50),(128,64),(150,75)]),
            'est__regressor__alpha': [1e-4,5e-4,1e-3,5e-3,1e-2],
            'est__regressor__learning_rate_init': [1e-3,3e-3,5e-3,1e-2],
            'est__regressor__batch_size': [32,64,128]
        }
        mlp = MLPRegressor(
            random_state=RANDOM_STATE, early_stopping=True, max_iter=800,
            n_iter_no_change=20, validation_fraction=0.15
        )
        search = RandomizedSearchCV(
            estimator=Pipeline([('ct', base_pipe), ('est', wrap_y(mlp))]),
            param_distributions=param_distributions, n_iter=n_trials, cv=tscv, scoring='r2',
            random_state=RANDOM_STATE, n_jobs=-1
        )
    elif name == 'SVM':
        param_distributions = {
            'est__regressor__kernel': ['rbf'],
            'est__regressor__C': [0.3,0.5,1,2,3],
            'est__regressor__gamma': [1e-4,3e-4,1e-3,3e-3,1e-2],
            'est__regressor__epsilon': [1e-3,3e-3,1e-2,3e-2,0.05]
        }
        svm = SVR()
        search = RandomizedSearchCV(
            estimator=Pipeline([('ct', base_pipe), ('est', wrap_y(svm))]),
            param_distributions=param_distributions, n_iter=n_trials, cv=tscv, scoring='r2',
            random_state=RANDOM_STATE, n_jobs=-1
        )
    elif name == 'RF':
        param_distributions = {
            'est__regressor__n_estimators': [100,150,200,250],
            'est__regressor__max_depth': [3,4,5,6,7],
            'est__regressor__min_samples_split': [10,15,20,25,30],
            'est__regressor__min_samples_leaf': [5,10,15,20],
            'est__regressor__max_features': ['sqrt']
        }
        search = RandomizedSearchCV(
            estimator=Pipeline([('ct', base_pipe), ('est', wrap_y(RandomForestRegressor(random_state=RANDOM_STATE, n_jobs=-1)))]),
            param_distributions=param_distributions, n_iter=n_trials, cv=tscv, scoring='r2',
            random_state=RANDOM_STATE, n_jobs=-1
        )
    else:  # XGB
        param_distributions = {
            'est__regressor__n_estimators': [150,200,250,300],
            'est__regressor__max_depth': [3,4,5],
            'est__regressor__learning_rate': [0.03,0.05,0.07,0.1],
            'est__regressor__subsample': [0.7,0.8,0.9,1.0],
            'est__regressor__colsample_bytree': [0.7,0.8,0.9,1.0],
            'est__regressor__reg_alpha': [0.1,0.3,1.0,3.0],
            'est__regressor__reg_lambda': [1.0,2.0,3.0,5.0],
        }
        xgb = XGBRegressor(objective='reg:squarederror', random_state=RANDOM_STATE, n_jobs=-1)
        search = RandomizedSearchCV(
            estimator=Pipeline([('ct', base_pipe), ('est', wrap_y(xgb))]),
            param_distributions=param_distributions, n_iter=n_trials, cv=tscv, scoring='r2',
            random_state=RANDOM_STATE, n_jobs=-1
        )

    search.fit(X, y)
    best_est = search.best_estimator_
    best_params = {k.replace('est__regressor__','').replace('est__',''):v for k,v in search.best_params_.items()}
    return best_est, best_params

# =============== Simple Stacking and Blending ===============
class StackingBaseLinear:
    def __init__(self, base_pipe):
        self.base_pipe = base_pipe
        self.meta = RidgeCV(alphas=np.logspace(-4,3,20))
    def fit(self, X, y):
        tscv = make_tscv(len(X))
        oof = np.full(len(X), np.nan)
        for tr,te in tscv.split(X):
            self.base_pipe.fit(X.iloc[tr], y.iloc[tr])
            oof[te] = self.base_pipe.predict(X.iloc[te])
        mask = ~np.isnan(oof)
        self.meta.fit(oof[mask].reshape(-1,1), y.iloc[mask])
        self.base_pipe.fit(X, y)
        return self
    def predict(self, X):
        base = self.base_pipe.predict(X)
        return self.meta.predict(base.reshape(-1,1))

class BlendTop2:
    def __init__(self, mdl1, mdl2, w=0.6):
        self.m1, self.m2, self.w = mdl1, mdl2, w
    def fit(self, X, y): return self
    def predict(self, X): return self.w*self.m1.predict(X) + (1-self.w)*self.m2.predict(X)

# =============== Forecasting (rolling + non-negative clipping + explosion protection) ===============
def iterative_forecast_tabular(best_model, hist_feat_df, future_feat_df, feats, target,
                               nonneg=True):
    last={}
    for city, g in hist_feat_df[hist_feat_df['Year']<=2024].groupby('City'):
        tail = g[target].dropna().tail(12).tolist()
        while len(tail)<12: tail=[0.0]+tail
        last[city]=tail[-12:]
    hist_cap = hist_feat_df.groupby('City')[target].max().to_dict()

    f = future_feat_df.copy()
    f['date'] = pd.to_datetime(dict(year=f['Year'], month=f['Month'], day=1))
    f.sort_values(['date','City'], inplace=True)
    rows=[]
    for (yr,mo), g in f.groupby(['Year','Month']):
        batch=[]; keys=[]
        for _, r in g.iterrows():
            city=r['City']
            if city not in last: last[city]=[0.0]*12
            seq=last[city]
            feat = {
                'people': float(r.get('people',0.0)),
                'total_GDP': float(r.get('total_GDP',0.0)),
                'Urbanization_rate': float(r.get('Urbanization_rate',0.0)),
                'Baidu_index': float(r.get('Baidu_index',0.0)),
                'PEV_Subsidy funds': float(r.get('PEV_Subsidy funds',0.0)),
                'month_sin': math.sin(2*math.pi*(int(mo)-1)/12),
                'month_cos': math.cos(2*math.pi*(int(mo)-1)/12),
                'year_num': int(yr),
                'lag_1': seq[-1], 'lag_3': seq[-3], 'lag_12': seq[-12],
                'roll_mean_3': np.mean(seq[-3:]),
                'roll_mean_12': np.mean(seq[-12:])
            }
            # Interaction and seasonal difference
            if 'gdp_per_capita' in feats and feat['people']>0:
                feat['gdp_per_capita'] = feat['total_GDP'] / (feat['people'] + 1e-9)
            if 'subsidy_to_gdp' in feats and feat['total_GDP']>0:
                feat['subsidy_to_gdp'] = feat['PEV_Subsidy funds'] / (feat['total_GDP'] + 1e-9)
            if 'baidu_per_capita' in feats and feat['people']>0:
                feat['baidu_per_capita'] = feat['Baidu_index'] / (feat['people'] + 1e-9)
            if 'sdiff_12' in feats:
                feat['sdiff_12'] = feat['lag_1'] - feat['lag_12']

            batch.append([feat.get(c,0.0) for c in feats]); keys.append(city)
        if not batch: continue
        Xb = pd.DataFrame(batch, columns=feats)
        yhat = best_model.predict(Xb)
        for city, val in zip(keys, yhat):
            v = float(val)
            if nonneg: v = max(0.0, v)
            cap = float(hist_cap.get(city, np.nan))
            # if np.isfinite(cap) and cap>0:
            #     v = min(v, cap*cap_mult)
            rows.append({'City':city,'Year':int(yr),'Month':int(mo), f'{target}_pred':v})
            last[city].append(v); last[city]=last[city][-12:]
    return pd.DataFrame(rows).sort_values(['City','Year','Month']).reset_index(drop=True)

# =============== Build "Locked Models" ===============
def build_locked_model(cluster_id, ct):
    info = LOCKED.get(cluster_id)
    if info is None: return None, None, None

    if info["type"] == "MLP":
        p = info["MLP_params"]
        mlp = MLPRegressor(
            hidden_layer_sizes=p["hidden_layer_sizes"],
            alpha=p["alpha"],
            learning_rate_init=p["learning_rate_init"],
            batch_size=p["batch_size"],
            early_stopping=True, max_iter=800,
            n_iter_no_change=20, validation_fraction=0.15,
            random_state=RANDOM_STATE
        )
        mdl = Pipeline([('ct', ct), ('est', wrap_y(mlp))])
        name = "ANN_MLP(locked)"
        best_params = {"type":"MLP", **p}
        return mdl, best_params, name

    if info["type"] == "BlendTop2":
        # Sub-model A: MLP
        pA = info["ANN_MLP_params"]
        mlp = MLPRegressor(
            hidden_layer_sizes=pA["hidden_layer_sizes"],
            alpha=pA["alpha"],
            learning_rate_init=pA["learning_rate_init"],
            batch_size=pA["batch_size"],
            early_stopping=True, max_iter=800,
            n_iter_no_change=20, validation_fraction=0.15,
            random_state=RANDOM_STATE
        )
        mdlA = Pipeline([('ct', ct), ('est', wrap_y(mlp))])

        # Sub-model B: SVM
        pB = info["SVM_params"]
        svm = SVR(kernel=pB["kernel"], C=pB["C"], gamma=pB["gamma"], epsilon=pB["epsilon"])
        mdlB = Pipeline([('ct', ct), ('est', wrap_y(svm))])

        blend = BlendTop2(mdlA, mdlB, w=info["blend_weight"])
        name = "BlendTop2(ANN_MLP+SVM)_locked"
        best_params = {"type":"BlendTop2", "blend_weight":info["blend_weight"],
                       "ANN_MLP_params":pA, "SVM_params":pB}
        return blend, best_params, name

    return None, None, None

# ========= Modified: Simplified Plot (no metrics display) =========
def plot_fit_train_test_onefig(train_df, test_df, feats, target, model, out_png,
                               title_prefix="", point_size=20):
    """
    Plot fitting scatter for training & test sets + marginal distributions on one figure
    """
    # Prediction
    y_tr_true = train_df[target].values
    y_tr_pred = model.predict(train_df[feats])
    y_te_true = test_df[target].values
    y_te_pred = model.predict(test_df[feats])

    # Unified coordinate range
    xy_min = min(y_tr_true.min(), y_te_true.min(), y_tr_pred.min(), y_te_pred.min())
    xy_max = max(y_tr_true.max(), y_te_true.max(), y_tr_pred.max(), y_te_pred.max())

    # JointGrid: scatter in middle, kernel density on edges
    g = sns.JointGrid(x=np.concatenate([y_tr_true, y_te_true]),
                      y=np.concatenate([y_tr_pred, y_te_pred]),
                      height=5.8)
    # Training points
    g.ax_joint.scatter(y_tr_true, y_tr_pred, s=point_size, alpha=0.8, label="Train", edgecolors='none')
    # Test points
    g.ax_joint.scatter(y_te_true, y_te_pred, s=point_size, alpha=0.9, label="Test", edgecolors='none')

    # 45-degree reference line
    g.ax_joint.plot([xy_min, xy_max], [xy_min, xy_max], linestyle='--')

    # Marginal distributions (only KDE for true and predicted values)
    sns.kdeplot(x=y_tr_true, ax=g.ax_marg_x, linewidth=1.5)
    sns.kdeplot(x=y_te_true, ax=g.ax_marg_x, linewidth=1.5)
    sns.kdeplot(y=y_tr_pred, ax=g.ax_marg_y, linewidth=1.5)
    sns.kdeplot(y=y_te_pred, ax=g.ax_marg_y, linewidth=1.5)

    g.ax_joint.set_xlabel("Actual")
    g.ax_joint.set_ylabel("Predicted")
    g.ax_joint.set_xlim(xy_min, xy_max)
    g.ax_joint.set_ylim(xy_min, xy_max)
    g.ax_joint.legend(loc="upper left")

    if title_prefix:
        g.ax_joint.set_title(title_prefix, pad=12)

    plt.tight_layout()
    g.figure.savefig(out_png, dpi=150)
    plt.close(g.figure)

# =============== Cluster-wise Training/Evaluation/Prediction ===============
def train_compare_cluster(sub_hist, sub_future, target, out_tag, cluster_id):
    sub_tab, feats = build_tabular(sub_hist.copy(), target)
    train_df = sub_tab[sub_tab['Year']<=2023]
    test_df  = sub_tab[sub_tab['Year']==2024]
    if test_df.empty and len(sub_tab)>=24:
        cutoff = sub_tab['date'].quantile(0.85)
        train_df = sub_tab[sub_tab['date']<=cutoff]
        test_df  = sub_tab[sub_tab['date']>cutoff]
    if len(train_df)==0 or len(test_df)==0:
        raise RuntimeError("Insufficient samples in this cluster")

    ct = make_feature_block(feats)

    # Priority: if cluster is in locked list, use locked model directly; otherwise perform limited search
    locked_model, locked_params, locked_name = build_locked_model(cluster_id, ct)
    MODELS = {}
    best_params_json = "{}"

    if locked_model is not None:
        # First fit sub-models for blend usage
        if isinstance(locked_model, BlendTop2):
            # Extract both sub-models
            locked_model.m1.fit(train_df[feats], train_df[target])
            locked_model.m2.fit(train_df[feats], train_df[target])
        else:
            locked_model.fit(train_df[feats], train_df[target])
        MODELS[locked_name] = locked_model
        best_params_json = json.dumps(locked_params, ensure_ascii=False)
    else:
        # Non-locked clusters: retain a small search set
        rf, rf_p   = fit_with_search('RF',  ct, train_df[feats], train_df[target])
        xgb, xgb_p = fit_with_search('XGB', ct, train_df[feats], train_df[target])
        mlp, mlp_p = fit_with_search('MLP', ct, train_df[feats], train_df[target])
        svm, svm_p = fit_with_search('SVM', ct, train_df[feats], train_df[target])
        MODELS = {
            'RF':rf, 'XGB':xgb, 'ANN_MLP':mlp, 'SVM':svm,
            'Stacking_MLP+Ridge': StackingBaseLinear(mlp).fit(train_df[feats], train_df[target]),
            'Stacking_SVM+Ridge': StackingBaseLinear(svm).fit(train_df[feats], train_df[target]),
        }

    # Evaluation
    eval_rows=[]
    for name, mdl in MODELS.items():
        y_pred = mdl.predict(test_df[feats])
        m = metrics_all(test_df[target], y_pred, train_df[target])
        m.update(model=name)
        eval_rows.append(m)
    eval_df = pd.DataFrame(eval_rows).sort_values(['R2','sMSE'], ascending=[False,True])

    # Report best
    best_name = eval_df.iloc[0]['model']
    best_model = MODELS[best_name]
    y_pred_tr = best_model.predict(train_df[feats])
    train_r2  = r2_score(train_df[target], y_pred_tr)
    test_r2   = eval_df.iloc[0]['R2']
    print(f"[{out_tag}] Best model: {best_name} | Test R2={test_r2:.3f}")

    # Plot - Only CityPlots will be generated
    city_plot_dir = os.path.join(OUT_DIR, f"{out_tag}_CityPlots"); os.makedirs(city_plot_dir, exist_ok=True)
    test_pred = test_df[['City','Year','Month',target]].copy()
    test_pred['pred'] = y_pred
    for city, g in test_pred.sort_values(['City','Year','Month']).groupby('City'):
        if len(g)==0: continue
        png_city = os.path.join(city_plot_dir, f"{out_tag}_{city}_2024_line.png")
        plt.figure(figsize=(6.5,3.2))
        plt.plot(g['Month'], g[target], marker='o', label='Actual')
        plt.plot(g['Month'], g['pred'],  marker='o', label='Predicted')
        plt.title(f"{out_tag} {city} - 2024")
        plt.xlabel("Month"); plt.ylabel(target); plt.legend()
        plt.tight_layout(); plt.savefig(png_city, dpi=150); plt.close()

    # <=2024 full retraining + future rolling prediction
    hist_tab = sub_tab[sub_tab['Year']<=2024].copy()
    if isinstance(best_model, BlendTop2):
        best_model.m1.fit(hist_tab[feats], hist_tab[target])
        best_model.m2.fit(hist_tab[feats], hist_tab[target])
    else:
        best_model.fit(hist_tab[feats], hist_tab[target])

    pred_future = iterative_forecast_tabular(best_model, hist_tab, sub_future.copy(), feats, target,
                                             nonneg=True)

    # Export evaluation table (including only required metrics)
    eval_out = eval_df.copy()
    eval_out.loc[eval_out['model']==best_name,'Train_R2'] = train_r2
    eval_out.loc[eval_out['model']==best_name,'Test_R2']  = test_r2
    eval_out.loc[eval_out['model']==best_name,'BestParams'] = best_params_json

    # Keep only required columns for evaluation output
    required_columns = ['model', 'R2', 'Train_R2', 'Test_R2', 'NMSE', 'BestParams']
    available_columns = [col for col in required_columns if col in eval_out.columns]
    eval_out_filtered = eval_out[available_columns]

    write_df(MASTER_XLSX, eval_out_filtered, sheet=f'{out_tag}_EvalAll', index=False)

    return best_name, best_params_json, pred_future, eval_df

# =============== Main Process (PEV only) ===============
def run_pev_only():
    df, t_pev = read_panel(INPUT_PATH, SHEET_NAME)

    # Light cleaning/imputation + mild winsorize (consistent with previous version)
    for col in ['people','total_GDP','Urbanization_rate','Baidu_index','PEV_Subsidy funds', t_pev]:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    for city, g in df.groupby('City'):
        idx = g.index
        df.loc[idx, [t_pev,'people','total_GDP','Urbanization_rate','Baidu_index','PEV_Subsidy funds']] = \
            g[[ 'PEV_number' if t_pev=='PEV_number' else t_pev,
                'people','total_GDP','Urbanization_rate','Baidu_index','PEV_Subsidy funds']].interpolate(limit_direction='both')
        s = df.loc[idx, t_pev].dropna()
        if len(s)>=20:
            lo, hi = np.percentile(s, 1), np.percentile(s, 99)
            df.loc[idx, t_pev] = df.loc[idx, t_pev].clip(lo, hi)

    hist = df[(df['Year']>=2016) & (df['Year']<=2024)].copy()
    fut  = df[(df['Year']>=2025) & (df['Year']<=2030)].copy()

    # Clustering
    print("\n==== [PEV] Auto K selection and clustering ====")
    city_feat = extract_city_level_features(hist, t_pev)
    emb = embed_city_features(city_feat)
    k_best, metric_tbl, score = choose_k_enhanced(emb)
    km = KMeans(n_clusters=k_best, n_init=100, random_state=RANDOM_STATE).fit(emb[['pc1','pc2']].values)
    labels = km.labels_
    emb_k = emb.assign(CityCluster=labels)

    # Write only required sheets to Excel
    write_df(MASTER_XLSX, emb_k.sort_values(['CityCluster','City']), sheet='PEV_Clusters', index=False)
    if metric_tbl is not None:
        mt = metric_tbl.copy()
        if score is not None: mt['score']=score.values
        write_df(MASTER_XLSX, mt, sheet='PEV_ClusterMetrics', index=False)
    print(f"→ [PEV] Best K = {k_best}, number of clusters: {k_best}")

    hist_k = hist.merge(emb_k[['City','CityCluster']], on='City', how='left')
    fut_k  = fut .merge(emb_k[['City','CityCluster']], on='City', how='left')

    # Cluster-wise processing
    best_rows, eval_concat, pred_concat = [], [], []
    for k in sorted(hist_k['CityCluster'].dropna().unique()):
        sub_hist = hist_k[hist_k['CityCluster']==k].copy()
        sub_future = fut_k[fut_k['CityCluster']==k].copy()
        if sub_hist.empty or sub_future.empty: continue
        print(f"  - Training cluster C{k} (number of cities={sub_hist['City'].nunique()})")
        try:
            best_name, best_params, pred_mon, eval_df = train_compare_cluster(
                sub_hist, sub_future, t_pev, out_tag=f'PEV_Cluster{k}', cluster_id=int(k))
            best_rows.append({'CityCluster':int(k),'BestModel':best_name,'BestParams':best_params,
                              'TestR2':float(eval_df.iloc[0]["R2"])})
            eval_concat.append(eval_df.assign(CityCluster=int(k)))
            pred_concat.append(pred_mon.assign(CityCluster=int(k)))
        except Exception as e:
            print(f"[WARN] Cluster C{k} failed:", e)
            traceback.print_exc()

    # Write only required evaluation results with filtered columns
    if eval_concat:
        E = pd.concat(eval_concat, ignore_index=True)
        # Keep only required columns
        required_columns = ['CityCluster', 'model', 'R2', 'Train_R2', 'NMSE']
        available_columns = [col for col in required_columns if col in E.columns]
        E_filtered = E[available_columns]
        write_df(MASTER_XLSX, E_filtered.sort_values(['CityCluster','R2'], ascending=[True,False]),
                 sheet='PEV_EvalAll', index=False)
    if best_rows:
        B = pd.DataFrame(best_rows)
        write_df(MASTER_XLSX, B.sort_values('CityCluster'), sheet='PEV_BestPerCluster', index=False)
    if pred_concat:
        P = pd.concat(pred_concat, ignore_index=True)
        write_df(MASTER_XLSX, P.sort_values(['City','Year','Month']), sheet='PEV_PredMonthly', index=False)
        Py = P.groupby(['City','Year'])[[f'{t_pev}_pred']].sum().reset_index()
        write_df(MASTER_XLSX, Py.sort_values(['City','Year']), sheet='PEV_PredAnnual', index=False)

    print(f"\n==== [PEV] Completed, results written to {MASTER_XLSX} ====\n")

if __name__ == "__main__":
    run_pev_only()
